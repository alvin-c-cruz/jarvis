from openpyxl import load_workbook
from dataclasses import dataclass
import os

from models import Accounts, Purchases, Supplier

PURCHASES_HEADERS = [
    'Receiving Date',
    'RR Number',
    'PO',
    'Invoice Number',
    'Particulars',
    'Invalid',
    'VAT Zero-Rated',
    'VAT Exempt',
    'VAT 12%',
    'EWT Rate',
    'Net of VAT',
]
SUPPLIER_HEADERS = [
    'Supplier Name',
    'TIN',
]

ACCOUNTS = {
    "Input VAT": {
        "account_number": "1501",
        "account_title": "Input Tax",
        "vat_class": "",
        "wt_class": "",
    },
    "EWT": {
        "account_number": "2201",
        "account_title": "Withholding Tax Payable - E",
        "vat_class": "",
        "wt_class": "",
    },
    "RAW MATS FOOD": {
        "account_number": "5001",
        "account_title": "Food Purchases",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "RAW MATS BEVERAGES": {
        "account_number": "5002",
        "account_title": "Beverage Purchases",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "BAR SUPPLIES": {
        "account_number": "6214",
        "account_title": "Bar Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "OFFICE SUPPLIES": {
        "account_number": "6212",
        "account_title": "Office Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "DINING SUPPLIES": {
        "account_number": "6218",
        "account_title": "Dining Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "GUEST SUPPLIES": {
        "account_number": "6217",
        "account_title": "Guest Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "CLEANING SUPPLIES": {
        "account_number": "6219",
        "account_title": "Cleaning Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "PACKAGING SUPPLIES": {
        "account_number": "6220",
        "account_title": "Packaging Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "MEDICAL SUPPLIES": {
        "account_number": "6229",
        "account_title": "Medical Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "UTENSILS / EQUIPMENT": {
        "account_number": "6211",
        "account_title": "Utensils Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "EM": {
        "account_number": "6109",
        "account_title": "Employees Meal",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "Insurance": {
        "account_number": "6308",
        "account_title": "Insurance",
        "vat_class": "Services",
        "wt_class": "2%",
    },
    "Accounting Fee": {
        "account_number": "6402",
        "account_title": "Accounting Expense",
        "vat_class": "Services",
        "wt_class": "10%",
    },
    "Security Services": {
        "account_number": "6313",
        "account_title": "Security Services",
        "vat_class": "Services",
        "wt_class": "2%",
    },
    "Pest Control": {
        "account_number": "6234",
        "account_title": "Pest Control Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "Marketing Support": {
        "account_number": "6315",
        "account_title": "Marketing Support",
        "vat_class": "Services",
        "wt_class": "2%",
    },
    "Consultancy": {
        "account_number": "6316",
        "account_title": "Consultancy Expense",
        "vat_class": "Services",
        "wt_class": "10%",
    },
    "Telephone": {
        "account_number": "6204",
        "account_title": "Telephone Expense",
        "vat_class": "Services",
        "wt_class": "2%",
    },
    "Others": {
        "account_number": "5101",
        "account_title": "Fuel and Gas Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "Accounts Payable": {
        "account_number": "2101",
        "account_title": "Accounts Payable",
        "vat_class": "",
        "wt_class": "",
    },

}


@dataclass
class RecordPurchases:
    db: any
    filename: str
    header_row: int = 4

    def __post_init__(self):
        wb = load_workbook(self.filename, data_only=True)

        for sheet_name in wb.sheetnames:
            self.process_sheet(wb[sheet_name])

    def process_sheet(self, ws):
        self.ws = ws

        self.column_headers = [self.ws.cell(row=self.header_row, column=column).value
                                for column in range(1, len(list(self.ws.columns)) + 1)
                           ]

        # Get saved account headers
        self.account_headers = [i[0] for i in self.db.execute("SELECT account_tag FROM tbl_accounts;").fetchall()]

        self.purchases_headers = PURCHASES_HEADERS
        self.supplier_headers = SUPPLIER_HEADERS

        self.process_column_headers()
        self.record_entries()

    def process_column_headers(self):
        choices = {str(header[0]+1): header[1] for header in enumerate(
            ('account_headers',
             'supplier_headers',
             'purchases_headers')
        )}

        for column_head in self.column_headers:
            if column_head not in self.account_headers + self.purchases_headers + self.supplier_headers \
                    and column_head is not None:
                header_type = ""
                print(choices)
                while header_type not in choices:
                    print(f"Header not in record: {column_head}")
                    header_type = input("Header type: \n") if column_head not in ACCOUNTS else "1"

                if choices[header_type] == 'account_headers':
                    # Saved account headers
                    self.account_headers.append(column_head)
                    account = Accounts(
                        db=self.db,
                        account_tag=column_head,
                        account_number=ACCOUNTS[column_head]['account_number'],
                        account_title=ACCOUNTS[column_head]['account_title'],
                        vat_class=ACCOUNTS[column_head]['vat_class'],
                        wt_class=ACCOUNTS[column_head]['wt_class'],
                    )
                    account.save
                else:
                    input(f"Header not in record: {column_head}")

                print()
                print()

    def record_entries(self):
        row_nums = range(self.header_row + 1, len(list(self.ws.rows))+1)
        columns = range(1, len(list(self.ws.columns)))

        receiving_date = str(self.ws.cell(row=self.header_row + 1, column=1).value)[:10]
        for row in row_nums:
            if self.ws.cell(row=row, column=2).value is None:
                break
            supplier = {}
            purchases = {}
            accounts = {}
            for column in columns:
                cell_value = self.ws.cell(row=row, column=column).value
                tag = self.ws.cell(row=self.header_row, column=column).value

                if tag in self.supplier_headers:
                    supplier[tag] = cell_value
                elif tag in self.purchases_headers:
                    if tag == 'Receiving Date':
                        if cell_value is None:
                            purchases[tag] = receiving_date
                        else:
                            purchases[tag] = receiving_date = str(cell_value)[:10]
                    else:
                        purchases[tag] = cell_value
                else:
                    if cell_value:
                        accounts[tag] = cell_value

            supplier_id = self.record_supplier(supplier)
            self.record_purchases(purchases, accounts, supplier_id)

    def record_supplier(self, supplier):
        supplier_name = supplier['Supplier Name']
        supplier_tin = supplier['TIN']

        if supplier_name:
            supplier = Supplier(db=self.db)
            if not self.db.execute('SELECT COUNT(*) FROM tbl_supplier WHERE supplier_name=?',
                                   (supplier_name,)).fetchone()[0]:
                supplier.supplier_name = supplier_name
                supplier.supplier_tin = supplier_tin
                supplier.save
                print("Added ", supplier)

            else:
                supplier.get(supplier_name=supplier_name)
                if not supplier.supplier_tin and supplier_tin:
                    supplier.supplier_tin = supplier_tin
                    supplier.save
                    print("Updated", supplier)

            return supplier.id

    def record_purchases(self, purchases, accounts, supplier_id):
        received_date = purchases['Receiving Date']
        rr_number = str(purchases['RR Number'])
        po = str(purchases['PO'])
        invoice = str(purchases['Invoice Number'])
        particulars = purchases['Particulars']
        invalid = purchases['Invalid'] if purchases['Invalid'] is not None else 0.0
        vat_zero_rated = purchases['VAT Zero-Rated'] if purchases['VAT Zero-Rated'] is not None else 0.0
        vat_exempt = purchases['VAT Exempt'] if purchases['VAT Exempt'] is not None else 0.0
        vat_12 = purchases['VAT 12%'] if purchases['VAT 12%'] is not None else 0.0
        ewt_rate = purchases['EWT Rate'] if purchases['EWT Rate'] is not None else 0.0

        tag = [account_tag for account_tag, value in accounts.items()
               if account_tag not in ('Input VAT', 'EWT', 'Accounts Payable')]

        if tag:
            account_id = self.db.execute('SELECT id FROM tbl_accounts WHERE account_tag=?',
                                                  (tag[0], )).fetchone()[0]
        else:
            account_id = 1

        purchase = Purchases(
            db=self.db,
            received_date=received_date,
            rr_number=rr_number,
            po=po,
            invoice=invoice,
            particulars=particulars,
            invalid=invalid,
            vat_zero_rated=vat_zero_rated,
            vat_exempt=vat_exempt,
            vat_12=vat_12,
            ewt_rate=ewt_rate,
            supplier_id=supplier_id,
            account_id=account_id,
        )

        purchase.save
        print(f"Added {purchase}")
