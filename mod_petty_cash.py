from openpyxl import load_workbook
from dataclasses import dataclass

from models import Accounts, PettyCash, Supplier

PETTY_CASH_HEADERS = [
    'Date',
    'PCV Number',
    'Invoice Number',
    'Particulars',
    'Invalid',
    'VAT Zero-Rated',
    'VAT Exempt',
    'VAT 12%',
    'EWT Rate',
    'Net of VAT',
]
SUPPLIER_HEADERS = [
    'Payee',
    'TIN',
    'Address',
]


ACCOUNTS = {
    "Input VAT": {
        "account_number": "1501",
        "account_title": "Input Tax",
        "vat_class": "",
        "wt_class": "",
    },
    "EWT": {
        "account_number": "2201",
        "account_title": "Withholding Tax Payable - E",
        "vat_class": "",
        "wt_class": "",
    },
    "RAW MATS FOOD": {
        "account_number": "5001",
        "account_title": "Food Purchases",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "RAW MATS BEVERAGES": {
        "account_number": "5002",
        "account_title": "Beverage Purchases",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "CLEANING ": {
        "account_number": "6219",
        "account_title": "Cleaning Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "CLEANING": {
        "account_number": "6219",
        "account_title": "Cleaning Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "PACKAGING": {
        "account_number": "6220",
        "account_title": "Packaging Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "OFFICE SUPPLIES": {
        "account_number": "6212",
        "account_title": "Office Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "DINING SUPPLIES": {
        "account_number": "6218",
        "account_title": "Dining Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "GUEST SUPPLIES": {
        "account_number": "6217",
        "account_title": "Guest Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "CLEANING SUPPLIES": {
        "account_number": "6219",
        "account_title": "Cleaning Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "KITCHEN SUPPLIES": {
        "account_number": "",
        "account_title": "Kitchen Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "DECORS": {
        "account_number": "",
        "account_title": "Decors Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "MEDICAL SUPPLIES": {
        "account_number": "6229",
        "account_title": "Medical Supplies Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "WARES AND UTENSILS": {
        "account_number": "6211",
        "account_title": "Utensils Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "REPAIRS AND MAINTENANCE": {
        "account_number": "",
        "account_title": "Repairs and Maintenance Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "PHOTOCOPY": {
        "account_number": "",
        "account_title": "Photocopy Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "TRANSPO": {
        "account_number": "",
        "account_title": "Transportation Expense",
        "vat_class": "",
        "wt_class": "",
    },
    "SALARIES AND WAGES": {
        "account_number": "",
        "account_title": "Salaries and Wages",
        "vat_class": "",
        "wt_class": "",
    },
    "MARKETING": {
        "account_number": "",
        "account_title": "Marketing Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "EMP MEAL": {
        "account_number": "6109",
        "account_title": "Employees Meal",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "MISC": {
        "account_number": "6999",
        "account_title": "Miscellaneous Expense",
        "vat_class": "Goods",
        "wt_class": "1%",
    },
    "Petty Cash": {
        "account_number": "",
        "account_title": "Petty Cash",
        "vat_class": "",
        "wt_class": "",
    },

}

@dataclass
class RecordPettyCash:
    db: any
    filename: str
    header_row: int = 4

    def __post_init__(self):
        wb = load_workbook(self.filename, data_only=True)

        for sheet_name in wb.sheetnames:
            self.process_sheet(wb[sheet_name])

    def process_sheet(self, ws):
        self.ws = ws

        self.column_headers = [self.ws.cell(row=self.header_row, column=column).value
                                for column in range(1, len(list(self.ws.columns)) + 1)
                           ]

        # Get saved account headers
        self.account_headers = [i[0] for i in self.db.execute("SELECT account_tag FROM tbl_accounts;").fetchall()]

        self.petty_cash_headers = PETTY_CASH_HEADERS
        self.supplier_headers = SUPPLIER_HEADERS

        self.process_column_headers()
        self.record_entries()

    def process_column_headers(self):
        choices = {str(header[0]+1): header[1] for header in enumerate(
            ('account_headers',
             'supplier_headers',
             'petty_cash_headers')
        )}

        for column_head in self.column_headers:
            if column_head not in self.account_headers + self.petty_cash_headers + self.supplier_headers \
                    and column_head is not None:
                header_type = ""
                print(choices)
                while header_type not in choices:
                    print(f"Header not in record: {column_head}")
                    header_type = input("Header type: \n") if column_head not in ACCOUNTS else "1"

                if choices[header_type] == 'account_headers':
                    # Saved account headers
                    self.account_headers.append(column_head)
                    if not self.db.execute('SELECT COUNT(*) FROM tbl_accounts WHERE account_title=?',
                                           (ACCOUNTS[column_head]['account_title'],)).fetchone()[0]:
                        account = Accounts(
                            db=self.db,
                            account_tag=column_head,
                            account_number=ACCOUNTS[column_head]['account_number'],
                            account_title=ACCOUNTS[column_head]['account_title'],
                            vat_class=ACCOUNTS[column_head]['vat_class'],
                            wt_class=ACCOUNTS[column_head]['wt_class'],
                        )
                        account.save
                        print(f'Added {account}')
                else:
                    input(f"Header not in record: {column_head}")

                print()
                print()

    def record_entries(self):
        row_nums = range(self.header_row + 1, len(list(self.ws.rows))+1)
        columns = range(1, len(list(self.ws.columns)))

        receiving_date = str(self.ws.cell(row=self.header_row + 1, column=1).value)[:10]
        for row in row_nums:
            if self.ws.cell(row=row, column=3).value is None:
                continue
            supplier = {}
            petty_cash = {}
            accounts = {}
            for column in columns:
                cell_value = self.ws.cell(row=row, column=column).value
                tag = self.ws.cell(row=self.header_row, column=column).value

                if tag in self.supplier_headers:
                    supplier[tag] = cell_value
                elif tag in self.petty_cash_headers:
                    if tag == 'Date':
                        if cell_value is None:
                            petty_cash[tag] = receiving_date
                        else:
                            petty_cash[tag] = receiving_date = str(cell_value)[ :10 ]
                    else:
                        petty_cash[tag] = cell_value
                else:
                    if cell_value:
                        accounts[tag] = cell_value

            supplier_id = self.record_supplier(supplier)
            self.record_petty_cash(petty_cash, accounts, supplier_id)

    def record_supplier(self, supplier):
        supplier_name = supplier['Payee']
        supplier_tin = supplier['TIN']

        if supplier_name:
            supplier = Supplier(db=self.db)
            if not self.db.execute('SELECT COUNT(*) FROM tbl_supplier WHERE supplier_name=?',
                                   (supplier_name,)).fetchone()[0]:
                supplier.supplier_name = supplier_name
                supplier.supplier_tin = supplier_tin
                supplier.save
                print("Added ", supplier)

            else:
                supplier.get(supplier_name=supplier_name)
                if not supplier.supplier_tin and supplier_tin:
                    supplier.supplier_tin = supplier_tin
                    supplier.save
                    print("Updated", supplier)

            return supplier.id

    def record_petty_cash(self, petty_cash, accounts, supplier_id):
        received_date = petty_cash['Date']
        invoice = str(petty_cash['Invoice Number'])
        particulars = petty_cash['Particulars']
        invalid = petty_cash['Invalid'] if petty_cash['Invalid'] is not None else 0.0
        vat_zero_rated = petty_cash['VAT Zero-Rated'] if petty_cash['VAT Zero-Rated'] is not None else 0.0
        vat_exempt = petty_cash['VAT Exempt'] if petty_cash['VAT Exempt'] is not None else 0.0
        vat_12 = petty_cash['VAT 12%'] if petty_cash['VAT 12%'] is not None else 0.0
        ewt_rate = petty_cash['EWT Rate'] if petty_cash['EWT Rate'] is not None else 0.0

        tag = [account_tag for account_tag, value in accounts.items()
               if account_tag not in ('Input VAT', 'EWT', 'Petty Cash')]

        if tag:
            account_id = self.db.execute('SELECT id FROM tbl_accounts WHERE account_tag=?',
                                                  (tag[0], )).fetchone()[0]
        else:
            account_id = 1

        pcf = PettyCash(
            db=self.db,
            received_date=received_date,
            invoice=invoice,
            particulars=particulars,
            invalid=invalid,
            vat_zero_rated=vat_zero_rated,
            vat_exempt=vat_exempt,
            vat_12=vat_12,
            ewt_rate=ewt_rate,
            supplier_id=supplier_id,
            account_id=account_id,
        )

        pcf.save
        print(f"Added {pcf}")
